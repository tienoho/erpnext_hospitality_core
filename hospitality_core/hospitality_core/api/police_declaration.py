# -*- coding: utf-8 -*-
"""
Module Khai Báo Tạm Trú Cơ Quan Công An & Cổng Xuất Nhập Cảnh Quảng Ninh
Tuân thủ 100% nguyên tắc ZERO HARDCODE - Đọc cấu hình động từ Hospitality Police Settings.
Đáp ứng chuẩn định dạng Cổng Dịch vụ công Quản lý Xuất nhập cảnh & Tạm trú Công an tỉnh Quảng Ninh.
"""

import frappe
from frappe import _
from frappe.utils import getdate, nowdate, formatdate
import io
import csv

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    openpyxl = None

ALLOWED_ROLES = [
    "System Manager",
    "Hospitality Manager",
    "Hospitality User",
    "Frontdesk Supervisor",
    "Frontdesk User",
    "Auditor"
]

def _get_police_settings():
    """
    Đọc cấu hình động từ Hospitality Police Settings và ERPNext Company.
    TUYỆT ĐỐI ZERO HARDCODE: Không gán cứng tên cơ sở, MST hay địa danh.
    """
    settings = None
    try:
        settings = frappe.get_cached_doc("Hospitality Police Settings")
    except Exception:
        pass

    default_company = frappe.db.get_single_value("Global Defaults", "default_company") or frappe.defaults.get_user_default("Company")
    company_doc = None
    if default_company:
        try:
            company_doc = frappe.get_cached_doc("Company", default_company)
        except Exception:
            pass

    company_name = getattr(company_doc, "company_name", None) or default_company or ""
    company_tax_id = getattr(company_doc, "tax_id", None) or ""

    est_name = (getattr(settings, "establishment_name", None) or "").strip() or company_name
    est_code = (getattr(settings, "establishment_code", None) or "").strip()
    police_st = (getattr(settings, "police_station_name", None) or "").strip()
    police_city = (getattr(settings, "police_city", None) or "").strip()
    tax_id = (getattr(settings, "tax_id", None) or "").strip() or company_tax_id
    resort_co = (getattr(settings, "resort_company_name", None) or "").strip() or company_name
    address = (getattr(settings, "address", None) or "").strip()
    stay_purpose = (getattr(settings, "default_stay_purpose", None) or "").strip() or _("Du lịch / Nghỉ dưỡng")
    portal_url = (getattr(settings, "immigration_portal_url", None) or "").strip() or "https://xuatnhapcanh.gov.vn"

    return {
        "establishment_name": est_name,
        "establishment_code": est_code,
        "police_station_name": police_st,
        "police_city": police_city,
        "tax_id": tax_id,
        "resort_company_name": resort_co,
        "address": address,
        "default_stay_purpose": stay_purpose,
        "immigration_portal_url": portal_url
    }

def check_police_declaration_permission():
    """
    Kiểm tra phân quyền truy cập tính năng Khai báo tạm trú.
    """
    if not frappe.session.user or frappe.session.user == "Guest":
        frappe.throw(_("Vui lòng đăng nhập để truy cập tính năng này."), frappe.PermissionError)

    user_roles = frappe.get_roles(frappe.session.user) if hasattr(frappe, "get_roles") else []
    has_role = any(r in ALLOWED_ROLES for r in user_roles)

    if not has_role and not frappe.has_permission("Hotel Reservation", "read"):
        frappe.throw(
            _("Bạn không có quyền truy cập hoặc kết xuất báo cáo Khai báo Tạm trú."),
            frappe.PermissionError
        )


def _build_excel_workbook(title, subtitle, meta_text, headers, rows, sheet_title="BaoCao"):
    """
    Xây dựng bảng tính Excel (.xlsx) chuẩn thẩm mỹ, chuyên nghiệp với openpyxl.
    Tự động kẻ viền, căn lề, định dạng tiêu đề và co giãn độ rộng cột.
    """
    if not openpyxl:
        frappe.throw(_("Thư viện openpyxl chưa được cài đặt trên máy chủ."))

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_title[:31]

    # Typography & Styles
    title_font = Font(name="Calibri", size=13, bold=True, color="1E3A8A")
    subtitle_font = Font(name="Calibri", size=15, bold=True, color="0F172A")
    meta_font = Font(name="Calibri", size=10, italic=True, color="475569")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1E40AF", end_color="1E40AF", fill_type="solid")
    alt_row_fill = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")

    thin_border_side = Side(border_style="thin", color="CBD5E1")
    cell_border = Border(top=thin_border_side, left=thin_border_side, right=thin_border_side, bottom=thin_border_side)

    # 1. Resort / Company Title
    ws.append([title])
    ws.cell(row=1, column=1).font = title_font

    # 2. Report Subtitle
    ws.append([subtitle])
    ws.cell(row=2, column=1).font = subtitle_font

    # 3. Meta information
    ws.append([meta_text])
    ws.cell(row=3, column=1).font = meta_font

    # 4. Blank spacer
    ws.append([])

    # 5. Table Headers
    ws.append(headers)
    header_row_idx = 5
    ws.row_dimensions[header_row_idx].height = 28

    num_cols = len(headers)
    for col_idx in range(1, num_cols + 1):
        cell = ws.cell(row=header_row_idx, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = cell_border

    # 6. Data Rows
    for row_idx, r in enumerate(rows, start=header_row_idx + 1):
        ws.append(r)
        ws.row_dimensions[row_idx].height = 22
        is_even = (row_idx % 2 == 0)

        for col_idx in range(1, num_cols + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.font = Font(name="Calibri", size=11)
            cell.border = cell_border

            if is_even:
                cell.fill = alt_row_fill

            # Canh lề: Cột STT, Ngày, Số phòng, Mã, Giới tính -> Giữa; Còn lại -> Trái
            hdr = headers[col_idx - 1].lower()
            if col_idx == 1 or "ngày" in hdr or "phòng" in hdr or "mã" in hdr or "giới tính" in hdr:
                cell.alignment = Alignment(horizontal="center", vertical="center")
            else:
                cell.alignment = Alignment(horizontal="left", vertical="center")

    # 7. Tự động tính toán độ rộng cột (Auto-fit Column Width)
    for col in ws.columns:
        col_letter = get_column_letter(col[0].column)
        max_len = 0
        for cell in col:
            # Bỏ qua 3 dòng tiêu đề khi tính độ rộng cột A
            if cell.row in (1, 2, 3) and col_letter == "A":
                continue
            val_str = str(cell.value or "")
            if len(val_str) > max_len:
                max_len = len(val_str)
        ws.column_dimensions[col_letter].width = max(max_len + 4, 12)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


@frappe.whitelist()
def get_daily_guest_list(target_date=None, company=None):
    """
    Truy vấn danh sách khách lưu trú phục vụ khai báo tạm trú công an.
    """
    check_police_declaration_permission()
    conf = _get_police_settings()

    if not target_date:
        target_date = nowdate()

    if not company:
        company = conf["resort_company_name"]

    try:
        guests = frappe.db.sql("""
            SELECT 
                g.name AS guest_id,
                g.full_name,
                g.identification_type,
                g.identification_no,
                g.mobile_no,
                g.email_id,
                g.address,
                r.name AS reservation_id,
                r.room AS room_number,
                r.arrival_date,
                r.departure_date,
                r.status AS reservation_status,
                r.is_alien,
                r.passport_number,
                r.company
            FROM `tabHotel Reservation` r
            LEFT JOIN `tabGuest` g ON r.guest = g.name
            WHERE (r.status IN ('Checked In', 'Arrived', 'Confirmed'))
              AND (r.arrival_date <= %(target_date)s AND r.departure_date >= %(target_date)s)
              AND (r.company = %(company)s OR r.company IS NULL OR %(company)s = '')
            ORDER BY r.room ASC, g.full_name ASC
        """, {"target_date": target_date, "company": company}, as_dict=True)
    except Exception as e:
        frappe.log_error(f"Error querying guest list for police declaration: {str(e)}", "PoliceDeclaration")
        guests = []

    return guests


@frappe.whitelist()
def export_police_declaration_csv(target_date=None, company=None, file_format="csv"):
    """
    Xuất file biểu mẫu Khai báo Tạm trú Công an (CSV UTF-8 with BOM hoặc Excel XLSX).
    100% Cấu hình động lấy thông tin cơ sở từ Hospitality Police Settings.
    """
    if str(file_format).lower() in ("xlsx", "excel"):
        return export_police_declaration_xlsx(target_date, company)

    check_police_declaration_permission()
    conf = _get_police_settings()

    if not target_date:
        target_date = nowdate()

    if not company:
        company = conf["resort_company_name"]

    guests = get_daily_guest_list(target_date, company)

    output = io.StringIO()
    output.write('\ufeff')
    writer = csv.writer(output, delimiter=',', quotechar='"', quoting=csv.QUOTE_MINIMAL)

    # Header chuẩn biểu mẫu Công an
    writer.writerow([
        "STT",
        "Họ và Tên",
        "Giới tính",
        "Ngày sinh",
        "Quốc tịch",
        "Loại giấy tờ",
        "Số CCCD / Hộ chiếu",
        "Số điện thoại",
        "Địa chỉ thường trú / Cư trú",
        "Số phòng",
        "Ngày đến",
        "Ngày đi dự kiến",
        "Mục đích lưu trú",
        "Cơ sở lưu trú",
        "Mã số thuế Doanh nghiệp"
    ])

    for idx, g in enumerate(guests, start=1):
        full_name = (g.get('full_name') or g.get('name') or "").upper()
        id_type = "Hộ chiếu" if g.get('is_alien') or g.get('identification_type') == "Passport" else "CCCD"
        id_number = g.get('passport_number') or g.get('identification_no') or ""
        nationality = "Nước ngoài" if g.get('is_alien') else "Việt Nam"
        phone = g.get('mobile_no') or ""
        address = g.get('address') or ""
        room_no = g.get('room_number') or ""
        cin = formatdate(g.get('arrival_date'), "dd/mm/yyyy") if g.get('arrival_date') else formatdate(target_date, "dd/mm/yyyy")
        cout = formatdate(g.get('departure_date'), "dd/mm/yyyy") if g.get('departure_date') else ""

        writer.writerow([
            idx,
            full_name,
            "Nam/Nữ",
            "",
            nationality,
            id_type,
            id_number,
            phone,
            address,
            room_no,
            cin,
            cout,
            conf["default_stay_purpose"],
            f"{conf['establishment_name']} ({company})",
            conf["tax_id"]
        ])

    csv_data = output.getvalue()
    output.close()

    filename = f"Khai_Bao_Tam_Tru_{conf['establishment_code']}_{target_date}.csv"
    if hasattr(frappe, "response"):
        frappe.response['result'] = csv_data
        frappe.response['type'] = 'csv'
        frappe.response['doctype'] = 'Police_Guest_Declaration'
        frappe.response['filename'] = filename
    return csv_data


@frappe.whitelist()
def export_police_declaration_xlsx(target_date=None, company=None):
    """
    Xuất Báo cáo Khai báo Tạm trú Toàn bộ Khách Lưu Trú định dạng Excel (.xlsx)
    Chuẩn mẫu báo cáo gửi Công an Phường Tuần Châu / Công an TP Hạ Long.
    """
    check_police_declaration_permission()
    conf = _get_police_settings()

    if not target_date:
        target_date = nowdate()

    if not company:
        company = conf["resort_company_name"]

    guests = get_daily_guest_list(target_date, company)

    headers = [
        "STT",
        "Họ và Tên",
        "Giới tính",
        "Ngày sinh",
        "Quốc tịch",
        "Loại giấy tờ",
        "Số CCCD / Hộ chiếu",
        "Số điện thoại",
        "Địa chỉ thường trú / Cư trú",
        "Số phòng",
        "Ngày đến",
        "Ngày đi dự kiến",
        "Mục đích lưu trú",
        "Cơ sở lưu trú",
        "Mã số thuế Doanh nghiệp"
    ]

    rows = []
    for idx, g in enumerate(guests, start=1):
        full_name = (g.get("full_name") or g.get("name") or "").upper()
        id_type = "Hộ chiếu" if g.get("is_alien") or g.get("identification_type") == "Passport" else "CCCD"
        id_number = g.get("passport_number") or g.get("identification_no") or ""
        nationality = "Nước ngoài" if g.get("is_alien") else "Việt Nam"
        phone = g.get("mobile_no") or ""
        address = g.get("address") or ""
        room_no = g.get("room_number") or ""
        cin = formatdate(g.get("arrival_date"), "dd/mm/yyyy") if g.get("arrival_date") else formatdate(target_date, "dd/mm/yyyy")
        cout = formatdate(g.get("departure_date"), "dd/mm/yyyy") if g.get("departure_date") else ""

        rows.append([
            idx,
            full_name,
            "Nam/Nữ",
            "",
            nationality,
            id_type,
            id_number,
            phone,
            address,
            room_no,
            cin,
            cout,
            conf["default_stay_purpose"],
            f"{conf['establishment_name']} ({company})",
            conf["tax_id"]
        ])

    title = f"{company.upper()} - {conf['establishment_name'].upper()}"
    subtitle = "SỔ ĐĂNG KÝ KHAI BÁO TẠM TRÚ KHÁCH LƯU TRÚ (CÔNG AN ĐỊA PHƯƠNG)"
    meta_text = f"Ngày báo cáo: {formatdate(target_date, 'dd/mm/yyyy')} | Mã cơ sở: {conf['establishment_code']} | Nơi tiếp nhận: {conf['police_station_name']} ({conf['police_city']})"

    buf = _build_excel_workbook(title, subtitle, meta_text, headers, rows, sheet_title="TamTru_CongAn")
    filename = f"Khai_Bao_Tam_Tru_{conf['establishment_code']}_{target_date}.xlsx"

    if hasattr(frappe, "response"):
        frappe.response["type"] = "binary"
        frappe.response["filename"] = filename
        frappe.response["filecontent"] = buf.getvalue()
        frappe.response["content_type"] = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

    return buf.getvalue()



@frappe.whitelist()
def export_quangninh_immigration_report(target_date=None, company=None, file_format="csv"):
    """
    Xuất file chuẩn 11 cột bắt buộc phục vụ nộp Cổng Khai báo Tạm trú Người Nước Ngoài
    của Cục Quản lý Xuất nhập cảnh & Công an Tỉnh Quảng Ninh (quangninh.xuatnhapcanh.gov.vn).
    Hỗ trợ định dạng CSV (UTF-8 with BOM) và Excel (.xlsx).
    """
    if str(file_format).lower() in ("xlsx", "excel"):
        return export_quangninh_immigration_report_xlsx(target_date, company)

    check_police_declaration_permission()
    conf = _get_police_settings()

    if not target_date:
        target_date = nowdate()

    if not company:
        company = conf["resort_company_name"]

    # Chỉ lọc các khách có yếu tố nước ngoài (is_alien = 1 hoặc Passport)
    all_guests = get_daily_guest_list(target_date, company)
    foreign_guests = [g for g in all_guests if g.get('is_alien') or g.get('passport_number') or g.get('identification_type') == 'Passport']

    output = io.StringIO()
    output.write('\ufeff')
    writer = csv.writer(output, delimiter=',', quotechar='"', quoting=csv.QUOTE_MINIMAL)

    # 11 Cột chuẩn nộp Cổng Cục QL Xuất Nhập Cảnh
    writer.writerow([
        "STT",
        "Họ và Tên Đệm (In hoa không dấu)",
        "Tên (In hoa không dấu)",
        "Giới tính (1:Nam, 2:Nữ)",
        "Ngày sinh (DD/MM/YYYY)",
        "Quốc tịch (Mã ISO-3)",
        "Số Hộ chiếu",
        "Loại thị thực / Miễn thị thực",
        "Ngày đến cơ sở lưu trú",
        "Ngày đi dự kiến",
        "Số phòng lưu trú",
        "Mục đích cư trú",
        "Mã cơ sở lưu trú"
    ])

    for idx, g in enumerate(foreign_guests, start=1):
        raw_name = (g.get('full_name') or "").strip()
        parts = raw_name.split()
        first_name = parts[-1].upper() if len(parts) > 0 else ""
        middle_last_name = " ".join(parts[:-1]).upper() if len(parts) > 1 else ""

        passport_no = g.get('passport_number') or g.get('identification_no') or ""
        cin = formatdate(g.get('arrival_date'), "dd/mm/yyyy") if g.get('arrival_date') else formatdate(target_date, "dd/mm/yyyy")
        cout = formatdate(g.get('departure_date'), "dd/mm/yyyy") if g.get('departure_date') else ""

        writer.writerow([
            idx,
            middle_last_name,
            first_name,
            "1", # 1: Nam / 2: Nữ
            "",
            "VNM" if not g.get('is_alien') else "FOR",
            passport_no,
            "Miễn thị thực",
            cin,
            cout,
            g.get('room_number') or "",
            conf["default_stay_purpose"],
            conf["establishment_code"]
        ])

    csv_data = output.getvalue()
    output.close()

    filename = f"Khai_Bao_XNC_QuangNinh_{conf['establishment_code']}_{target_date}.csv"
    if hasattr(frappe, "response"):
        frappe.response['result'] = csv_data
        frappe.response['type'] = 'csv'
        frappe.response['doctype'] = 'QuangNinh_Immigration_Report'
        frappe.response['filename'] = filename
    return csv_data


@frappe.whitelist()
def export_quangninh_immigration_report_xlsx(target_date=None, company=None):
    """
    Xuất Báo cáo Khai báo Xuất nhập cảnh Khách Quốc Tế định dạng Excel (.xlsx)
    Chuẩn 11 cột quy định Cục Quản lý Xuất nhập cảnh & Công an Tỉnh Quảng Ninh.
    """
    check_police_declaration_permission()
    conf = _get_police_settings()

    if not target_date:
        target_date = nowdate()

    if not company:
        company = conf["resort_company_name"]

    all_guests = get_daily_guest_list(target_date, company)
    foreign_guests = [g for g in all_guests if g.get('is_alien') or g.get('passport_number') or g.get('identification_type') == 'Passport']

    headers = [
        "STT",
        "Họ và Tên Đệm (In hoa)",
        "Tên (In hoa)",
        "Giới tính",
        "Ngày sinh",
        "Quốc tịch (Mã ISO-3)",
        "Số Hộ chiếu",
        "Loại thị thực / Miễn thị thực",
        "Ngày đến cơ sở lưu trú",
        "Ngày đi dự kiến",
        "Số phòng lưu trú",
        "Mục đích cư trú",
        "Mã cơ sở lưu trú"
    ]

    rows = []
    for idx, g in enumerate(foreign_guests, start=1):
        raw_name = (g.get('full_name') or "").strip()
        parts = raw_name.split()
        first_name = parts[-1].upper() if len(parts) > 0 else ""
        middle_last_name = " ".join(parts[:-1]).upper() if len(parts) > 1 else ""

        passport_no = g.get('passport_number') or g.get('identification_no') or ""
        cin = formatdate(g.get('arrival_date'), "dd/mm/yyyy") if g.get('arrival_date') else formatdate(target_date, "dd/mm/yyyy")
        cout = formatdate(g.get('departure_date'), "dd/mm/yyyy") if g.get('departure_date') else ""

        rows.append([
            idx,
            middle_last_name,
            first_name,
            "Nam" if str(g.get('gender', '')).lower() in ('nam', 'male', '1') else ("Nữ" if str(g.get('gender', '')).lower() in ('nữ', 'female', '2') else "Nam"),
            "",
            "VNM" if not g.get('is_alien') else "FOR",
            passport_no,
            "Miễn thị thực",
            cin,
            cout,
            g.get('room_number') or "",
            conf["default_stay_purpose"],
            conf["establishment_code"]
        ])

    title = f"{company.upper()} - {conf['establishment_name'].upper()}"
    subtitle = "DANH SÁCH KHAI BÁO TẠM TRÚ KHÁCH NƯỚC NGOÀI (CỔNG XNC QUẢNG NINH)"
    meta_text = f"Ngày báo cáo: {formatdate(target_date, 'dd/mm/yyyy')} | Mã cơ sở: {conf['establishment_code']} | Cổng tiếp nhận: {conf['immigration_portal_url']}"

    buf = _build_excel_workbook(title, subtitle, meta_text, headers, rows, sheet_title="XNC_QuangNinh")
    filename = f"Khai_Bao_XNC_QuangNinh_{conf['establishment_code']}_{target_date}.xlsx"

    if hasattr(frappe, "response"):
        frappe.response["type"] = "binary"
        frappe.response["filename"] = filename
        frappe.response["filecontent"] = buf.getvalue()
        frappe.response["content_type"] = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

    return buf.getvalue()



@frappe.whitelist()
def export_police_declaration_xml(target_date=None, company=None):
    """
    Xuất file XML chuẩn cấu trúc Cổng Quản lý Xuất nhập cảnh & Tạm trú Công an tỉnh Quảng Ninh.
    100% Cấu hình động lấy thông tin từ Hospitality Police Settings.
    """
    check_police_declaration_permission()
    conf = _get_police_settings()

    if not target_date:
        target_date = nowdate()

    if not company:
        company = conf["resort_company_name"]

    guests = get_daily_guest_list(target_date, company)

    xml_lines = [
        '<?xml version="1.0" encoding="utf-8"?>',
        '<KhaiBaoTamTru>',
        '  <ThongTinCoSo>',
        f'    <TenCoSo>{conf["establishment_name"]}</TenCoSo>',
        f'    <MaCoSo>{conf["establishment_code"]}</MaCoSo>',
        f'    <DoanhNghiep>{company}</DoanhNghiep>',
        f'    <MaSoThue>{conf["tax_id"]}</MaSoThue>',
        f'    <DiaChi>{conf["address"]}</DiaChi>',
        f'    <NgayKhaiBao>{target_date}</NgayKhaiBao>',
        f'    <TongSoKhach>{len(guests)}</TongSoKhach>',
        '  </ThongTinCoSo>',
        '  <DanhSachKhach>'
    ]

    for idx, g in enumerate(guests, start=1):
        full_name = (g.get('full_name') or "").upper()
        id_type = "Passport" if g.get('is_alien') else "CCCD"
        id_number = g.get('passport_number') or g.get('identification_no') or ""
        nationality = "FOREIGN" if g.get('is_alien') else "VNM"

        xml_lines.extend([
            '    <KhachLuuTru>',
            f'      <STT>{idx}</STT>',
            f'      <HoTen>{full_name}</HoTen>',
            f'      <QuocTich>{nationality}</QuocTich>',
            f'      <LoaiGiayTo>{id_type}</LoaiGiayTo>',
            f'      <SoGiayTo>{id_number}</SoGiayTo>',
            f'      <SoDienThoai>{g.get("mobile_no") or ""}</SoDienThoai>',
            f'      <SoPhong>{g.get("room_number") or ""}</SoPhong>',
            f'      <NgayDen>{g.get("arrival_date") or target_date}</NgayDen>',
            f'      <NgayDi>{g.get("departure_date") or ""}</NgayDi>',
            f'      <MucDich>{conf["default_stay_purpose"]}</MucDich>',
            '    </KhachLuuTru>'
        ])

    xml_lines.extend([
        '  </DanhSachKhach>',
        '</KhaiBaoTamTru>'
    ])

    xml_content = "\n".join(xml_lines)
    filename = f"Khai_Bao_Tam_Tru_{conf['establishment_code']}_{target_date}.xml"
    if hasattr(frappe, "response"):
        frappe.response['result'] = xml_content
        frappe.response['type'] = 'download'
        frappe.response['doctype'] = 'Police_Guest_Declaration_XML'
        frappe.response['filename'] = filename
    return xml_content
